import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
wb = openpyxl.load_workbook(r'C:\Users\Yoav\source\repos\MO6\sitedocs_internet_accessibility_form.xlsx', data_only=True)
for name in wb.sheetnames:
    print('=== SHEET:', name, '===')
    ws = wb[name]
    for row in ws.iter_rows(values_only=False):
        any_val = False
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip():
                any_val = True
                print(f'{cell.coordinate}: {v}')
        if any_val:
            print('---')
